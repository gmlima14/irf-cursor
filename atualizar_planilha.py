import pandas as pd
from openpyxl import load_workbook
from glob import glob
import os
import re
import xlsxwriter
from datetime import datetime

# Caminhos fixos
caminho_origem = r'C:\Users\CSUGAB01\Downloads'
caminho_destino = r'S:\Procurement\FUP\OTP Mensal\OTP - Base.xlsx'
nome_aba_destino = 'Base OTP'

# Função para extrair número dos 14 primeiros dígitos do nome do arquivo
def extrair_numero(nome_arquivo):
    match = re.match(r"(\d{14})", nome_arquivo)
    if match:
        return int(match.group(1))
    return -1

def log_message(message):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"[{timestamp}] {message}")

log_message('🚀 Início do processo de atualização de planilha.')

# Busca arquivos de origem com 'CELONIS' no nome
arquivos_origem = glob(os.path.join(caminho_origem, '*CELONIS*.csv'))
log_message(f'📁 Quantidade de arquivos de origem CELONIS encontrados: {len(arquivos_origem)}')

maior_numero = -1
arquivo_origem = None
for arquivo in arquivos_origem:
    nome = os.path.basename(arquivo)
    numero = extrair_numero(nome)
    if numero > maior_numero:
        maior_numero = numero
        arquivo_origem = arquivo

if not arquivo_origem:
    log_message("❌ Nenhum arquivo de origem CSV encontrado com 'CELONIS' e 14 dígitos no início do nome.")
    exit(1)
log_message(f'✅ Arquivo de origem CELONIS selecionado: {arquivo_origem}')

# Lê os dados do CSV usando pandas
try:
    df = pd.read_csv(arquivo_origem, encoding='utf-8')
except Exception as e:
    log_message(f'❌ Erro ao ler o arquivo CSV de origem: {e}')
    exit(1)
log_message(f'📊 {len(df)} linhas copiadas da planilha de origem.')

# Busca arquivos de origem com 'EXPORT_' no nome
arquivos_export = glob(os.path.join(caminho_origem, '*EXPORT_*.xlsx'))
log_message(f'📁 Quantidade de arquivos de origem EXPORT_ encontrados: {len(arquivos_export)}')

# Função para extrair número da data dos últimos 15 caracteres
def extrair_numero_export(nome_arquivo):
    # Formato esperado: EXPORT_20250709_134041.xlsx
    # Extrair YYYYMMDD_HHMMSS
    if 'EXPORT_' in nome_arquivo:
        # Remove a extensão .xlsx
        nome_sem_ext = nome_arquivo.replace('.xlsx', '')
        # Pega a parte após EXPORT_
        parte_data = nome_sem_ext.split('EXPORT_')[1]
        # Remove o '_' entre data e hora
        numero_str = parte_data.replace('_', '')
        try:
            return int(numero_str)
        except ValueError:
            return -1
    return -1

maior_numero_export = -1
arquivo_export = None
for arquivo in arquivos_export:
    nome = os.path.basename(arquivo)
    numero = extrair_numero_export(nome)
    if numero > maior_numero_export:
        maior_numero_export = numero
        arquivo_export = arquivo

if not arquivo_export:
    log_message("❌ Nenhum arquivo de origem Excel encontrado com 'EXPORT_' e data válida no final do nome.")
else:
    log_message(f'✅ Arquivo de origem EXPORT selecionado: {arquivo_export}')
    
    # Lê os dados do arquivo Excel EXPORT usando pandas
    try:
        df_export = pd.read_excel(arquivo_export)
    except Exception as e:
        log_message(f'❌ Erro ao ler o arquivo Excel EXPORT de origem: {e}')
    else:
        log_message(f'📊 {len(df_export)} linhas copiadas do arquivo EXPORT.')

# Carrega a planilha de destino para leitura da aba Base Fornecedores
wb_destino = load_workbook(caminho_destino, data_only=True, read_only=True)

# --- LÓGICA PARA ATUALIZAR 'Base Fornecedores' ---
aba_fornecedores = 'Base Fornecedores'
coluna_sap = 'SAP-LIFNR'
coluna_vendor = 'Vendor'

# Lista para armazenar novos fornecedores
novos_fornecedores = []

# Verifica se a aba existe
if aba_fornecedores not in wb_destino.sheetnames:
    log_message(f"⚠️ Aba '{aba_fornecedores}' inexistente na planilha de destino.")
    valores_sap = set()
else:
    ws_forn = wb_destino[aba_fornecedores]

    # Lê todos os valores existentes na coluna 'SAP-LIFNR' (ignorando cabeçalho)
    valores_sap = set()
    for row in ws_forn.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
        if row[0] is not None:
            valores_sap.add(str(row[0]))

    # Filtra valores únicos da coluna 'Vendor' do DataFrame de origem
    if coluna_vendor not in df.columns:
        log_message(f"❌ Coluna '{coluna_vendor}' não encontrada no arquivo de origem!")
    else:
        novos_vendors = set(df[coluna_vendor].dropna().astype(str).unique())
        novos_a_adicionar = novos_vendors - valores_sap
        log_message(f'🔍 {len(novos_a_adicionar)} novos fornecedores a adicionar.')
        
        # Armazena os novos fornecedores para adicionar depois
        if novos_a_adicionar:
            novos_fornecedores = sorted(novos_a_adicionar)

wb_destino.close()

# Lê todas as abas existentes para preservar
wb_existente = load_workbook(caminho_destino, data_only=True, read_only=True)
abas_existentes = {}
for sheet_name in wb_existente.sheetnames:
    if sheet_name not in [nome_aba_destino, 'RNC Base']:  # Não preservar as abas que vamos sobrescrever
        ws_temp = wb_existente[sheet_name]
        dados_aba = []
        for row in ws_temp.iter_rows(values_only=True):
            dados_aba.append(list(row))
        abas_existentes[sheet_name] = dados_aba
wb_existente.close()

# Cria novo arquivo com xlsxwriter para escrita otimizada
log_message('⚡ Criando novo arquivo com xlsxwriter...')
workbook = xlsxwriter.Workbook(caminho_destino, {'nan_inf_to_errors': True})

# Cria formato de data
date_format = workbook.add_format({'num_format': 'dd/mm/yyyy'})

# Define a data de hoje para usar na aba RNC Base
data_hoje = datetime.today().strftime('%d/%m/%Y')

# Cria aba Base OTP (primeira aba)
worksheet_otp = workbook.add_worksheet(nome_aba_destino)

# Escreve cabeçalho
for col, header in enumerate(df.columns):
    worksheet_otp.write(0, col, header)

# Converte colunas de data de uma vez
colunas_data = ['BEDAT', 'Due Date (incl. ex works time)', 'GR Document Date', 'Delivery Date']
for coluna in colunas_data:
    if coluna in df.columns:
        df[coluna] = pd.to_datetime(df[coluna], errors='coerce')

# Escreve dados em lote
for row_idx, row in enumerate(df.itertuples(index=False, name=None), start=1):
    for col_idx, value in enumerate(row):
        # Trata valores nulos/NaN
        if pd.isna(value):
            worksheet_otp.write(row_idx, col_idx, '')
        else:
            # Verifica se é uma coluna de data
            col_name = df.columns[col_idx]
            if col_name in colunas_data:
                # Já convertido para datetime, aplica formato
                if pd.notna(value):
                    worksheet_otp.write_datetime(row_idx, col_idx, value, date_format)
                else:
                    worksheet_otp.write(row_idx, col_idx, '')
            else:
                worksheet_otp.write(row_idx, col_idx, value)

log_message('📋 Dados atualizados na planilha Base OTP.')

# Cria aba RNC Base como segunda aba se houver dados do EXPORT
if arquivo_export and 'df_export' in locals():
    aba_rnc = 'RNC Base'
    worksheet_rnc = workbook.add_worksheet(aba_rnc)
    
    # Adiciona a coluna "Última Atualização" ao cabeçalho
    col_ultima_atualizacao_rnc = len(df_export.columns)
    worksheet_rnc.write(0, col_ultima_atualizacao_rnc, "Última Atualização")
    
    # Escreve a data de hoje apenas na primeira linha da nova coluna
    worksheet_rnc.write(1, col_ultima_atualizacao_rnc, data_hoje)
    
    # Escreve cabeçalho
    for col, header in enumerate(df_export.columns):
        worksheet_rnc.write(0, col, header)
    
    # Converte coluna de data de uma vez
    if 'Notification Date' in df_export.columns:
        df_export['Notification Date'] = pd.to_datetime(df_export['Notification Date'], errors='coerce')
    
    # Adiciona "0" no início da coluna Supplier se começar com número
    if 'Supplier' in df_export.columns:
        df_export['Supplier'] = df_export['Supplier'].astype(str).apply(
            lambda x: '0' + x if x and x[0].isdigit() else x
        )
    
    # Substitui valores NaN por string vazia na coluna 'Assembly Descript.'
    if 'Assembly Descript.' in df_export.columns:
        df_export['Assembly Descript.'] = df_export['Assembly Descript.'].fillna('')
    
    # Escreve dados em lote
    for row_idx, row in enumerate(df_export.itertuples(index=False, name=None), start=1):
        for col_idx, value in enumerate(row):
            # Trata valores nulos/NaN
            if pd.isna(value):
                worksheet_rnc.write(row_idx, col_idx, '')
            else:
                # Verifica se é uma coluna de data
                col_name = df_export.columns[col_idx]
                if col_name == 'Notification Date':
                    # Já convertido para datetime, aplica formato
                    if pd.notna(value):
                        worksheet_rnc.write_datetime(row_idx, col_idx, value, date_format)
                    else:
                        worksheet_rnc.write(row_idx, col_idx, '')
                else:
                    worksheet_rnc.write(row_idx, col_idx, value)
    
    log_message('📋 Dados atualizados na planilha RNC Base.')

# Recria as abas existentes preservadas
for nome_aba, dados_aba in abas_existentes.items():
    worksheet = workbook.add_worksheet(nome_aba)
    for row_idx, row_data in enumerate(dados_aba):
        for col_idx, value in enumerate(row_data):
            if pd.isna(value):
                worksheet.write(row_idx, col_idx, '')
            else:
                worksheet.write(row_idx, col_idx, value)

# Adiciona novos fornecedores na aba Base Fornecedores se existir
if novos_fornecedores and aba_fornecedores in abas_existentes:
    # Encontra a aba Base Fornecedores no workbook
    worksheet_forn = None
    for sheet in workbook.worksheets():
        if sheet.name == aba_fornecedores:
            worksheet_forn = sheet
            break
    
    if worksheet_forn:
        # Adiciona os novos fornecedores ao final
        start_row = len(abas_existentes[aba_fornecedores])
        for i, fornecedor in enumerate(novos_fornecedores):
            worksheet_forn.write_string(start_row + i, 0, fornecedor)

# Fecha o workbook
workbook.close()
log_message(f"🎉 Dados atualizados em: {caminho_destino}")
